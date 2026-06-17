#!/usr/bin/env python3
"""
Ingestor das peças da LDO / AMF (Anexo de Metas Fiscais) em formato Excel.
================================================================
Lê os anexos padronizados (layout LRF, art. 4º) já estruturados em Excel pela
equipe (Emerson) e popula as tabelas de planejamento do Radar 360:

  Anexo I  (Metas Fiscais Analítico) → ldo_metas_fiscais  (receita/despesa 3 anos, %RCL)
  Anexo V  (Quadro de Programas e Ações):
     aba "Resumo Programas" → orgaos + programas (+ total_estimado)
     aba "Ações"            → acoes (+ custo_estimado, meta_fisica)

Idempotente: apaga o que já existe do (cod_ibge, exercicio) desta fonte antes de
reinserir. Generaliza p/ qualquer município no padrão — os anexos seguem a LRF.

Uso:
  python3 ingest_ldo_amf.py --cod 3530706 --exercicio 2027 \\
      --anexo1 "/caminho/Anexo I - Metas Fiscais Analítico.xlsx" \\
      --anexo5 "/caminho/Anexo V - Quadro Geral de  Programas e Ações - Substituto.xlsx"
"""
import argparse
import os
import sys

try:
    import openpyxl
    import psycopg2
except ImportError as e:
    print(f"ERRO: dependência faltando ({e}). pip install openpyxl psycopg2-binary")
    sys.exit(1)

BASE = os.path.dirname(os.path.abspath(__file__))


def db_url():
    url = os.environ.get("DATABASE_URL")
    if url:
        return url
    for line in open(os.path.join(BASE, "..", ".env.local")):
        if line.startswith("DATABASE_URL="):
            return line.split("=", 1)[1].strip().strip('"').strip("'")
    raise SystemExit("DATABASE_URL não encontrado")


def num(x):
    if x is None:
        return None
    try:
        return float(x)
    except (TypeError, ValueError):
        return None


# Linhas do Anexo I que viram indicadores de meta fiscal. (rótulo no Excel → indicador)
AMF_INDICADORES = {
    "Receita Total (EXCETO FONTES RPPS)": "receita_total",
    "Receitas Primárias (EXCETO FONTES RPPS) (I)": "receita_primaria",
    "Despesa Total (EXCETO FONTES RPPS)": "despesa_total",
    "Despesas Primárias (EXCETO FONTES RPPS) (II)": "despesa_primaria",
    "Pessoal e Encargos Sociais": "pessoal_encargos",
    "Outras Despesas Correntes": "outras_despesas_correntes",
    "Reserva de Contingência": "reserva_contingencia",
}


def parse_anexo1(path):
    """→ list[(exercicio, indicador, valor_corrente, pct_rcl)]. Colunas por ano:
    Valor Corrente(a) e %RCL(a/RCL) ficam em offsets fixos a cada bloco de 4."""
    wb = openpyxl.load_workbook(path, data_only=True)
    out = []
    # ordem dos exercícios vem do cabeçalho "ESPECIFICAÇÃO | 2027 | 2028 | 2029"
    for sheet in ("Receitas", "Despesas"):
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        anos = []
        for r in rows:
            if r and r[0] and str(r[0]).strip().upper() == "ESPECIFICAÇÃO":
                anos = [int(c) for c in r[1:] if isinstance(c, (int, float))]
                break
        if not anos:
            continue
        for r in rows:
            rot = (str(r[0]).strip() if r and r[0] else "")
            if rot in AMF_INDICADORES:
                ind = AMF_INDICADORES[rot]
                # cada ano = bloco de 4 colunas: corrente, constante, %PIB, %RCL
                for i, ano in enumerate(anos):
                    base = 1 + i * 4
                    valor = num(r[base]) if len(r) > base else None
                    pct_rcl = num(r[base + 3]) if len(r) > base + 3 else None
                    if valor:
                        out.append((ano, ind, valor, pct_rcl))
    # Dedupe (ano, indicador) — o rótulo reaparece na seção "Composição".
    # A tabela principal vem antes, então a 1ª ocorrência é a correta.
    seen = set()
    deduped = []
    for ano, ind, valor, pct in out:
        if (ano, ind) in seen:
            continue
        seen.add((ano, ind))
        deduped.append((ano, ind, valor, pct))
    return deduped


def parse_anexo5(path):
    """→ (orgaos:set, programas:list, acoes:list)."""
    wb = openpyxl.load_workbook(path, data_only=True)
    orgaos = []
    programas = []
    acoes = []

    ws = wb["Resumo Programas"]
    prog_by_cod = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or not r[4]:  # Código Programa
            continue
        orgao = str(r[1]).strip() if r[1] else "—"
        unidade = str(r[3]).strip() if r[3] else None
        codigo = str(r[4]).strip()
        nome = str(r[5]).strip() if r[5] else ""
        objetivo = str(r[6]).strip() if r[6] else None
        total = num(r[8])
        if orgao not in orgaos:
            orgaos.append(orgao)
        # Mesmo código pode reaparecer em unidades distintas do órgão — agrega
        # o total e mantém a 1ª descrição (órgão/nome/objetivo).
        if codigo in prog_by_cod:
            if total:
                prog_by_cod[codigo]["total"] = (prog_by_cod[codigo]["total"] or 0) + total
        else:
            prog_by_cod[codigo] = {"orgao": orgao, "unidade": unidade, "codigo": codigo,
                                   "nome": nome, "objetivo": objetivo, "total": total}
    programas = list(prog_by_cod.values())

    if "Ações" in wb.sheetnames:
        ws = wb["Ações"]
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r or not r[3]:  # Cód. Ação
                continue
            acoes.append({
                "prog_codigo": str(r[1]).strip() if r[1] else "",
                "codigo": str(r[3]).strip(),
                "nome": str(r[4]).strip() if r[4] else "",
                "produto": str(r[6]).strip() if r[6] else None,
                "unidade": str(r[7]).strip() if r[7] else None,
                "meta_fisica": str(r[8]).strip() if r[8] is not None else None,
                "custo": num(r[9]),
            })
    return orgaos, programas, acoes


def ensure_columns(cur):
    """Aditivo e idempotente — nunca dropa."""
    cur.execute("ALTER TABLE programas ADD COLUMN IF NOT EXISTS total_estimado numeric")
    cur.execute("ALTER TABLE programas ADD COLUMN IF NOT EXISTS unidade text")
    cur.execute("ALTER TABLE acoes ADD COLUMN IF NOT EXISTS custo_estimado numeric")
    cur.execute("ALTER TABLE acoes ADD COLUMN IF NOT EXISTS meta_fisica text")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--cod", type=int, required=True)
    ap.add_argument("--exercicio", type=int, required=True)
    ap.add_argument("--anexo1", required=True)
    ap.add_argument("--anexo5", required=True)
    a = ap.parse_args()

    metas = parse_anexo1(a.anexo1)
    orgaos, programas, acoes = parse_anexo5(a.anexo5)
    print(f"Parse: {len(metas)} metas fiscais, {len(orgaos)} órgãos, "
          f"{len(programas)} programas, {len(acoes)} ações")

    conn = psycopg2.connect(db_url())
    cur = conn.cursor()
    ensure_columns(cur)

    # --- ldo_metas_fiscais (apaga e reinsere p/ este município) ---
    cur.execute("DELETE FROM ldo_metas_fiscais WHERE cod_ibge=%s AND exercicio = ANY(%s)",
                (a.cod, sorted({m[0] for m in metas})))
    for ano, ind, valor, pct in metas:
        cur.execute(
            "INSERT INTO ldo_metas_fiscais (cod_ibge, exercicio, indicador, meta_valor, meta_pct, base_legal) "
            "VALUES (%s,%s,%s,%s,%s,%s) "
            "ON CONFLICT (cod_ibge, exercicio, indicador) DO UPDATE "
            "SET meta_valor=EXCLUDED.meta_valor, meta_pct=EXCLUDED.meta_pct, base_legal=EXCLUDED.base_legal",
            (a.cod, ano, ind, valor, pct, "LDO AMF Demonstrativo I (LRF art. 4º, §1º)"),
        )

    # --- orgaos (upsert por nome) ---
    orgao_id = {}
    for nome in orgaos:
        cur.execute("SELECT id FROM orgaos WHERE cod_ibge=%s AND nome=%s", (a.cod, nome))
        row = cur.fetchone()
        if row:
            orgao_id[nome] = row[0]
        else:
            cur.execute("INSERT INTO orgaos (cod_ibge, nome, tipo, ativo) VALUES (%s,%s,'secretaria',true) RETURNING id",
                        (a.cod, nome))
            orgao_id[nome] = cur.fetchone()[0]

    # --- programas (apaga do exercício e reinsere) ---
    cur.execute("DELETE FROM acoes WHERE programa_id IN (SELECT id FROM programas WHERE cod_ibge=%s AND exercicio=%s)",
                (a.cod, a.exercicio))
    cur.execute("DELETE FROM programas WHERE cod_ibge=%s AND exercicio=%s", (a.cod, a.exercicio))
    prog_id = {}
    for p in programas:
        cur.execute(
            "INSERT INTO programas (cod_ibge, exercicio, codigo, nome, objetivo, unidade, total_estimado, orgao_id) "
            "VALUES (%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id",
            (a.cod, a.exercicio, p["codigo"], p["nome"], p["objetivo"], p["unidade"], p["total"], orgao_id.get(p["orgao"])),
        )
        prog_id[p["codigo"]] = cur.fetchone()[0]

    # --- acoes ---
    n_ac = 0
    for ac in acoes:
        pid = prog_id.get(ac["prog_codigo"])
        if not pid:
            continue
        cur.execute(
            "INSERT INTO acoes (programa_id, codigo, nome, produto, unidade_medida, meta_fisica, custo_estimado) "
            "VALUES (%s,%s,%s,%s,%s,%s,%s) "
            "ON CONFLICT (programa_id, codigo) DO UPDATE "
            "SET custo_estimado = COALESCE(acoes.custo_estimado,0) + COALESCE(EXCLUDED.custo_estimado,0)",
            (pid, ac["codigo"], ac["nome"], ac["produto"], ac["unidade"], ac["meta_fisica"], ac["custo"]),
        )
        n_ac += 1

    conn.commit()
    print(f"OK: {len(metas)} metas, {len(orgao_id)} órgãos, {len(prog_id)} programas, {n_ac} ações "
          f"gravados p/ cod_ibge={a.cod}, exercicio={a.exercicio}")
    conn.close()


if __name__ == "__main__":
    main()
