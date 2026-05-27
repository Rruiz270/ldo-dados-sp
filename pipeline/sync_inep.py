#!/usr/bin/env python3
"""
INEP → Neon DB sync
====================
Lê os XLSX baixados por inep_scraper.py e popula:
  - ideb                  (cod_ibge, rede, etapa, ciclo_avaliacao, ...)
  - indicadores_externos  (fonte_id='INEP', categoria='educacao',
                           indicador in 'taxa_aprovacao_*',
                           'taxa_reprovacao_*', 'taxa_abandono_*')

Filtro SP: só persistimos municípios com cod_ibge entre 3500000-3599999.

Estratégia IDEB:
  Cada arquivo IDEB_<ano>.xlsx traz HISTÓRICO de IDEB (várias colunas
  VL_OBSERVADO_<ciclo>). O arquivo 2023 cobre 2017, 2019, 2021, 2023 — para
  AI/AF inclui projecao para 2017+. Para EM, 2017 só tem observado no 2023
  (projeção pra EM começou no IDEB 2017). Para ter 2017 projecao no AI/AF
  caímos no arquivo 2019.

Ciclos 2025+: não estão disponíveis (Saeb 2025 ainda em campo).

Uso:
  export DATABASE_URL="postgresql://..."   # ou auto-detect via .env.local
  python3 sync_inep.py
"""

import csv
import json
import os
import re
import sys
import zipfile
from collections import defaultdict
from datetime import date, datetime
from io import BytesIO

try:
    import openpyxl
except ImportError:
    print("ERRO: openpyxl não instalado. Rode: python3 -m pip install --user openpyxl")
    sys.exit(1)

try:
    import psycopg2
    import psycopg2.extras
except ImportError:
    print("ERRO: psycopg2 não instalado. Rode: python3 -m pip install --user psycopg2-binary")
    sys.exit(1)

BASE = os.path.dirname(os.path.abspath(__file__))
INEP_DIR = os.path.join(BASE, "inep_data")

DATABASE_URL = os.environ.get("DATABASE_URL")
if not DATABASE_URL:
    envpath = "/Users/raphaelruiz/Projects/ldo-dados-sp/.env.local"
    if os.path.exists(envpath):
        with open(envpath) as f:
            for line in f:
                if line.startswith("DATABASE_URL="):
                    DATABASE_URL = line.split("=", 1)[1].strip().strip('"').strip("'")
                    break

# --- escopo ---
SP_COD_MIN = 3500000
SP_COD_MAX = 3599999

IDEB_YEARS_TO_LOAD = [2019, 2021, 2023]  # arquivos a abrir (cada um cobre histórico)
ETAPAS = ["anos_iniciais", "anos_finais", "ensino_medio"]
CICLOS_DESEJADOS = [2017, 2019, 2021, 2023]  # 2025 ainda não saiu


def log(msg):
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}", flush=True)


# --------------------------------------------------------------------
# Normalização
# --------------------------------------------------------------------

def norm_rede(s):
    """Normaliza coluna Rede pra valores estáveis."""
    if not s:
        return None
    v = str(s).strip().lower()
    # remove acentos básicos
    v = v.replace("ú", "u").replace("ó", "o").replace("á", "a") \
         .replace("é", "e").replace("ê", "e").replace("í", "i")
    if v.startswith("estadu"):  return "estadual"
    if v.startswith("municip"): return "municipal"
    if v.startswith("federal"): return "federal"
    if v.startswith("priv"):    return "privada"
    if v.startswith("pub"):     return "publica"
    return v


def norm_dep_admin(s):
    """Para o dataset de taxa de rendimento (NO_DEPENDENCIA)."""
    if not s:
        return None
    v = str(s).strip().lower()
    v = v.replace("ú", "u").replace("ó", "o").replace("á", "a") \
         .replace("é", "e").replace("ê", "e").replace("í", "i")
    if v == "total": return "total"
    if v.startswith("estadu"):  return "estadual"
    if v.startswith("municip"): return "municipal"
    if v.startswith("federal"): return "federal"
    if v.startswith("priv"):    return "privada"
    if v.startswith("pub"):     return "publica"
    return v


def to_float(v):
    """'-' / '*' / None / NaN -> None. Floats e strings numéricas viram float."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        if isinstance(v, float) and (v != v):  # NaN
            return None
        return float(v)
    s = str(v).strip().replace(",", ".")
    if s in ("", "-", "*", "**", "ND", "—"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


# --------------------------------------------------------------------
# Parser IDEB
# --------------------------------------------------------------------

ETAPA_FILE_RE = re.compile(r"ideb_(anos_iniciais|anos_finais|ensino_medio)_municipios_(\d{4})\.xlsx")


def parse_ideb_xlsx(path, etapa):
    """Itera linhas de um arquivo IDEB e devolve dicts.

    Retorna iterator de dicts:
      {cod_ibge, rede, ciclo,
       ideb_observado, ideb_projetado, nota_mat, nota_lp, fluxo}

    Detecta a linha de cabeçalho (busca por SG_UF na coluna 0).
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    header = None
    col = {}  # col_name -> idx
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        # detectar cabeçalho
        if header is None:
            if row and row[0] == "SG_UF":
                header = list(row)
                for idx, h in enumerate(header):
                    if h:
                        col[str(h)] = idx
            continue

        cod_raw = row[col.get("CO_MUNICIPIO", 1)]
        if cod_raw is None or not isinstance(cod_raw, (int, float)):
            continue
        try:
            cod = int(cod_raw)
        except (TypeError, ValueError):
            continue
        if cod < SP_COD_MIN or cod > SP_COD_MAX:
            continue

        rede = norm_rede(row[col.get("REDE", 3)])
        if not rede:
            continue

        # Para cada ciclo desejado, monta um registro se houver dado.
        for ciclo in CICLOS_DESEJADOS:
            obs_col = f"VL_OBSERVADO_{ciclo}"
            proj_col = f"VL_PROJECAO_{ciclo}"
            mat_col = f"VL_NOTA_MATEMATICA_{ciclo}"
            lp_col  = f"VL_NOTA_PORTUGUES_{ciclo}"
            rend_col = f"VL_INDICADOR_REND_{ciclo}"

            obs = to_float(row[col[obs_col]]) if obs_col in col else None
            proj = to_float(row[col[proj_col]]) if proj_col in col else None
            mat = to_float(row[col[mat_col]]) if mat_col in col else None
            lp  = to_float(row[col[lp_col]])  if lp_col in col else None
            rend = to_float(row[col[rend_col]]) if rend_col in col else None

            # Se NENHUM dado, pula
            if obs is None and proj is None and mat is None and lp is None and rend is None:
                continue

            yield {
                "cod_ibge": cod,
                "rede": rede,
                "etapa": etapa,
                "ciclo_avaliacao": ciclo,
                "ideb_observado": obs,
                "ideb_projetado": proj,
                "nota_padronizada_lp": lp,
                "nota_padronizada_mat": mat,
                "fluxo": rend,
            }


def build_ideb_rows():
    """Combina os 3 arquivos (2019, 2021, 2023) com preferência pelo mais
    recente (geralmente traz dados mais completos do histórico)."""
    # Estratégia: começamos pelo mais antigo e sobrescrevemos com o mais novo.
    by_key = {}  # (cod, rede, etapa, ciclo) -> dict
    for year_file in IDEB_YEARS_TO_LOAD:
        for etapa in ETAPAS:
            path = os.path.join(
                INEP_DIR, "ideb", str(year_file),
                f"ideb_{etapa}_municipios_{year_file}.xlsx",
            )
            if not os.path.exists(path):
                log(f"  [pular] não existe: {path}")
                continue
            log(f"  parsing IDEB {year_file}/{etapa}")
            count = 0
            for rec in parse_ideb_xlsx(path, etapa):
                key = (rec["cod_ibge"], rec["rede"], rec["etapa"],
                       rec["ciclo_avaliacao"])
                # Merge: preferimos valores não-NULL do arquivo mais novo,
                # mas mantemos do antigo se o novo for NULL.
                existing = by_key.get(key)
                if existing:
                    for field in ("ideb_observado", "ideb_projetado",
                                  "nota_padronizada_lp", "nota_padronizada_mat",
                                  "fluxo"):
                        if rec[field] is not None:
                            existing[field] = rec[field]
                else:
                    by_key[key] = rec
                count += 1
            log(f"    {count} registros candidatos lidos")
    log(f"  IDEB total únicos: {len(by_key)} (cod_ibge, rede, etapa, ciclo)")
    return list(by_key.values())


# --------------------------------------------------------------------
# Parser Taxas de Rendimento
# --------------------------------------------------------------------

# Mapeamento prefixo numérico -> tipo de taxa
TAXA_PREFIXES = {
    "1": "taxa_aprovacao",
    "2": "taxa_reprovacao",
    "3": "taxa_abandono",
}

# Colunas que queremos extrair (sufixos no header da planilha tx_rend):
# Total fundamental ({prefix}_CAT_FUN), anos iniciais ({prefix}_CAT_FUN_AI),
# anos finais ({prefix}_CAT_FUN_AF), ensino médio ({prefix}_CAT_MED).
NIVEL_SUFFIXES = {
    "CAT_FUN":    "ef_total",
    "CAT_FUN_AI": "ef_ai",
    "CAT_FUN_AF": "ef_af",
    "CAT_MED":    "em_total",
}


def parse_tx_rend_zip(zip_path, ano_censo):
    """Lê o XLSX dentro do ZIP de taxa de rendimento e yield dicts.

    Schema do INEP variou no tempo:
      - 2020+: tem linha técnica `NU_ANO_CENSO|NO_REGIAO|...|1_CAT_FUN|...`
      - 2019:  só tem cabeçalho humano (linha 5: "Ano|Região|UF|Código..."),
               sem nomes técnicos. Caímos pra mapeamento posicional.
    """
    with zipfile.ZipFile(zip_path) as z:
        xlsx_name = next(
            (n for n in z.namelist() if n.endswith(".xlsx")), None
        )
        if not xlsx_name:
            log(f"  WARN: nenhum XLSX dentro de {zip_path}")
            return
        data = z.read(xlsx_name)

    wb = openpyxl.load_workbook(BytesIO(data), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    header = None
    col = {}
    # Layout 2019: linhas de cabeçalho merged + dados começam na primeira linha
    # com inteiro na col 0 (ano censo). Total 65 colunas, sem header técnico.
    # Layout 2019 mapeado posicionalmente:
    LEGACY_COLS = {
        "CO_MUNICIPIO": 3,
        "NO_CATEGORIA": 5,
        "NO_DEPENDENCIA": 6,
        "1_CAT_FUN": 7,  "1_CAT_FUN_AI": 8,  "1_CAT_FUN_AF": 9,  "1_CAT_MED": 19,
        "2_CAT_FUN": 25, "2_CAT_FUN_AI": 26, "2_CAT_FUN_AF": 27, "2_CAT_MED": 37,
        "3_CAT_FUN": 43, "3_CAT_FUN_AI": 44, "3_CAT_FUN_AF": 45, "3_CAT_MED": 55,
    }

    legacy_mode = False
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if header is None:
            # Modo moderno: cabeçalho técnico explícito
            if row and row[0] == "NU_ANO_CENSO":
                header = list(row)
                for idx, h in enumerate(header):
                    if h:
                        col[str(h)] = idx
                continue
            # Modo legacy (2019): primeira linha de dados tem int na col 0
            if row and isinstance(row[0], int) and row[0] == ano_censo:
                header = []  # marker que entramos no modo legacy
                col = LEGACY_COLS
                legacy_mode = True
                # cai pro processamento abaixo nesta mesma linha
            else:
                continue

        cod_raw = row[col.get("CO_MUNICIPIO", 3)]
        try:
            cod = int(cod_raw)
        except (TypeError, ValueError):
            continue
        if cod < SP_COD_MIN or cod > SP_COD_MAX:
            continue

        categoria = (row[col.get("NO_CATEGORIA", 5)] or "").strip()
        # NO_CATEGORIA = Total/Urbana/Rural. Filtramos só "Total" pra evitar
        # ambiguidade na chave UNIQUE.
        if categoria.lower() != "total":
            continue

        dep = norm_dep_admin(row[col.get("NO_DEPENDENCIA", 6)])
        if not dep:
            continue

        # Extrai valor por (taxa, nivel)
        for prefix, taxa in TAXA_PREFIXES.items():
            for suffix, nivel in NIVEL_SUFFIXES.items():
                col_name = f"{prefix}_{suffix}"
                if col_name not in col:
                    continue
                v = to_float(row[col[col_name]])
                if v is None:
                    continue
                yield {
                    "cod_ibge": cod,
                    "dep": dep,
                    "ano_censo": ano_censo,
                    "indicador": f"{taxa}_{nivel}_{dep}",
                    "valor": v,
                }


def build_tx_rend_rows():
    """Itera todos os ZIPs disponíveis e gera linhas pra indicadores_externos."""
    rows = []
    for year in sorted(os.listdir(os.path.join(INEP_DIR, "tx_rend"))) \
            if os.path.exists(os.path.join(INEP_DIR, "tx_rend")) else []:
        m = re.match(r"tx_rend_municipios_(\d{4})\.zip", year)
        if not m:
            continue
        ano = int(m.group(1))
        path = os.path.join(INEP_DIR, "tx_rend", year)
        log(f"  parsing Tx Rendimento {ano}")
        count = 0
        for rec in parse_tx_rend_zip(path, ano):
            rows.append((
                rec["cod_ibge"],
                "INEP",
                rec["indicador"],
                "educacao",
                date(ano, 12, 31),
                rec["valor"],
                None,
                "pct",
                json.dumps({"ano_censo": ano, "dep_admin": rec["dep"]}),
            ))
            count += 1
        log(f"    {count} (taxa × nivel × dep) lidos")
    log(f"  Tx Rendimento total: {len(rows)} linhas")
    return rows


# --------------------------------------------------------------------
# UPSERTs
# --------------------------------------------------------------------

def upsert_ideb(conn, recs):
    if not recs:
        return
    rows = [
        (r["cod_ibge"], r["rede"], r["etapa"], r["ciclo_avaliacao"],
         r["ideb_observado"], r["ideb_projetado"],
         r["nota_padronizada_lp"], r["nota_padronizada_mat"],
         r["fluxo"], "INEP")
        for r in recs
    ]
    with conn.cursor() as cur:
        psycopg2.extras.execute_values(cur, """
            INSERT INTO ideb (
              cod_ibge, rede, etapa, ciclo_avaliacao,
              ideb_observado, ideb_projetado,
              nota_padronizada_lp, nota_padronizada_mat,
              fluxo, fonte_id
            )
            VALUES %s
            ON CONFLICT (cod_ibge, rede, etapa, ciclo_avaliacao) DO UPDATE SET
              ideb_observado = EXCLUDED.ideb_observado,
              ideb_projetado = EXCLUDED.ideb_projetado,
              nota_padronizada_lp = EXCLUDED.nota_padronizada_lp,
              nota_padronizada_mat = EXCLUDED.nota_padronizada_mat,
              fluxo = EXCLUDED.fluxo,
              fonte_id = EXCLUDED.fonte_id,
              atualizado_em = NOW()
        """, rows, page_size=500)
    log(f"  ideb: upserted {len(rows)} linhas")


def upsert_indicadores_externos(conn, rows):
    """rows: (cod_ibge, fonte_id, indicador, categoria, periodo_referencia,
             valor_numerico, valor_texto, unidade, metadata_json)"""
    if not rows:
        return
    with conn.cursor() as cur:
        psycopg2.extras.execute_values(cur, """
            INSERT INTO indicadores_externos (
              cod_ibge, fonte_id, indicador, categoria, periodo_referencia,
              valor_numerico, valor_texto, unidade, metadata
            )
            VALUES %s
            ON CONFLICT (cod_ibge, fonte_id, indicador, periodo_referencia) DO UPDATE SET
              categoria = EXCLUDED.categoria,
              valor_numerico = EXCLUDED.valor_numerico,
              valor_texto = EXCLUDED.valor_texto,
              unidade = EXCLUDED.unidade,
              metadata = EXCLUDED.metadata,
              atualizado_em = NOW()
        """, rows, page_size=1000)
    log(f"  indicadores_externos: upserted {len(rows)} linhas")


# --------------------------------------------------------------------
# CLI: dry-run mode (sem conexão Neon) p/ teste local
# --------------------------------------------------------------------

def dry_run_preview(ideb_recs, tx_rows, sample_codigos):
    """Mostra dados parseados para um set de municípios sem tocar no DB."""
    print()
    print("=" * 70)
    print("DRY-RUN: prévia de dados IDEB para municípios sample")
    print("=" * 70)
    for cod in sample_codigos:
        subset = [r for r in ideb_recs if r["cod_ibge"] == cod]
        if not subset:
            print(f"\n  {cod}: nenhum dado")
            continue
        print(f"\n  cod_ibge={cod} ({len(subset)} registros):")
        for r in sorted(subset, key=lambda x: (x["etapa"], x["rede"], x["ciclo_avaliacao"])):
            print(
                f"    {r['etapa']:13s} {r['rede']:10s} ciclo={r['ciclo_avaliacao']}: "
                f"IDEB obs={r['ideb_observado']} proj={r['ideb_projetado']} "
                f"LP={r['nota_padronizada_lp']} MAT={r['nota_padronizada_mat']} "
                f"fluxo={r['fluxo']}"
            )
    print()
    print(f"  Tx Rendimento sample (primeiras 6 do município {sample_codigos[0]}):")
    sample_tx = [r for r in tx_rows if r[0] == sample_codigos[0]][:6]
    for r in sample_tx:
        print(f"    {r[2]:50s} {r[4]} = {r[5]}{r[7] or ''}")


def main():
    start = datetime.now()
    log("=" * 64)
    log("INEP → Neon sync")
    log("=" * 64)

    dry_run = "--dry-run" in sys.argv

    log("Construindo IDEB rows a partir dos XLSX...")
    ideb_recs = build_ideb_rows()
    log(f"  IDEB SP: {len(ideb_recs)} registros únicos")

    log("Construindo Taxa de Rendimento rows...")
    tx_rows = build_tx_rend_rows()

    if dry_run:
        # Sample: Adamantina (3500105), São Paulo capital (3550308), Campinas (3509502)
        dry_run_preview(ideb_recs, tx_rows, [3550308, 3509502, 3500105])
        log(f"DRY-RUN ok em {datetime.now() - start}")
        return

    if not DATABASE_URL:
        log("ERRO: DATABASE_URL não definido (e --dry-run não passado)")
        sys.exit(1)

    log("Conectando ao Neon...")
    conn = psycopg2.connect(DATABASE_URL)
    conn.autocommit = False

    try:
        log("Upserting ideb...")
        upsert_ideb(conn, ideb_recs)
        log("Upserting indicadores_externos (tx_rendimento)...")
        upsert_indicadores_externos(conn, tx_rows)
        conn.commit()
        log(f"Commit OK em {datetime.now() - start}")
    except Exception as e:
        conn.rollback()
        log(f"FALHA: {e}")
        raise
    finally:
        conn.close()


if __name__ == "__main__":
    main()
